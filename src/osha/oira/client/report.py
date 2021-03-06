from Acquisition import aq_inner
from AccessControl import getSecurityManager
from cStringIO import StringIO
from collections import defaultdict
from datetime import datetime
from euphorie.client import model
from euphorie.client import report
from euphorie.client import survey
from euphorie.client.session import SessionManager
from euphorie.client import config
from euphorie.content.profilequestion import IProfileQuestion
from euphorie.content.interfaces import ICustomRisksModule
from euphorie.ghost import PathGhost
from five import grok
from openpyxl.cell import get_column_letter
from openpyxl.workbook import Workbook
from osha.oira import _
from osha.oira.client import utils
from osha.oira.client.interfaces import IOSHAIdentificationPhaseSkinLayer
from osha.oira.client.interfaces import IOSHAReportPhaseSkinLayer
from osha.oira.client.survey import OSHAStatus
from plonetheme.nuplone.utils import formatDate
from rtfng.Elements import PAGE_NUMBER
from rtfng.PropertySets import BorderPropertySet
from rtfng.PropertySets import FramePropertySet
from rtfng.PropertySets import ParagraphPropertySet
from rtfng.PropertySets import TabPropertySet
from rtfng.PropertySets import TextPropertySet
from rtfng.Renderer import Renderer
from rtfng.Styles import ParagraphStyle
from rtfng.Styles import TextStyle
from rtfng.document import character
from rtfng.document.base import TAB, LINE
from rtfng.document.paragraph import Paragraph, Table, Cell
from rtfng.document.section import Section
from sqlalchemy import sql
from z3c.saconfig import Session
from zExceptions import NotFound
from zope.i18n import translate
from zope.i18nmessageid import MessageFactory
import htmllaundry
import logging
import urllib


PloneLocalesFactory = MessageFactory("plonelocales")

log = logging.getLogger(__name__)

grok.templatedir("templates")


class ReportView(report.ReportView):
    """ Override so that skipped or filled in company forms are not shown
    again. For #4436.
    """
    grok.layer(IOSHAReportPhaseSkinLayer)
    grok.template("report")

    def update(self):
        self.session = SessionManager.session

        if self.request.environ["REQUEST_METHOD"] == "POST":
            self.session.report_comment = self.request.form.get("comment")

            url = "%s/report/company" % self.request.survey.absolute_url()
            if getattr(self.session, 'company', None) is not None and \
                    getattr(self.session.company, 'country') is not None:
                url = "%s/report/view" % self.request.survey.absolute_url()

            user = getSecurityManager().getUser()
            if getattr(user, 'account_type', None) == config.GUEST_ACCOUNT:
                url = "%s/@@register?report_blurb=1&came_from=%s" % (
                    self.request.survey.absolute_url(),
                    urllib.quote(url, '')
                )
            self.request.response.redirect(url)
            return


COLUMN_ORDER = [
    ('risk', 'title'),
    ('risk', 'priority'),
    ('measure', 'action_plan'),
    ('measure', 'prevention_plan'),
    ('measure', 'requirements'),
    ('measure', 'planning_end'),
    ('measure', 'responsible'),
    ('measure', 'budget'),
    ('risk', 'number'),
    ('module', 'title'),
    ('risk', 'comment'),
]


class ReportLanding(grok.View):
    """Custom report landing page.

    This replaces the standard online view of the report with a page
    offering the RTF and XLSX download options.
    """
    grok.context(PathGhost)
    grok.require("euphorie.client.ViewSurvey")
    grok.layer(IOSHAReportPhaseSkinLayer)
    grok.template("report_landing")
    grok.name("view")


class ActionPlanTimeline(report.ActionPlanTimeline):
    grok.layer(IOSHAReportPhaseSkinLayer)

    combine_keys = ['prevention_plan', 'requirements']
    columns = sorted(
        (col for col in report.ActionPlanTimeline.columns
            if (col[0], col[1]) in COLUMN_ORDER),
        key=lambda d, co=COLUMN_ORDER: co.index((d[0], d[1])))
    extra_cols = [x for x in columns if x[1] in combine_keys]
    columns = [x for x in columns if x[1] not in combine_keys]
    columns[2] = (
        'measure', 'action_plan',
        _('report_timeline_measure', default=u'Measure'))
    columns[3] = (
        'measure', 'planning_end',
        _('report_timeline_end_date', default=u'End date'))
    columns[4] = (
        'measure', 'responsible',
        _('report_timeline_responsible', default=u'Responsible'))
    columns.insert(-1, (None, None, _(
        'report_timeline_progress',
        default=u'Status (planned, in process, implemented)')))

    def create_workbook(self):
        """Create an Excel workbook containing the all risks and measures.
        """
        t = lambda txt: translate(txt, context=self.request)
        book = Workbook()
        sheet = book.worksheets[0]
        sheet.title = t(_('report_timeline_title', default=u'Timeline'))
        sheet.default_column_dimension.auto_size = True

        for (column, (type, key, title)) in enumerate(self.columns):
            if key in self.combine_keys:
                continue
            cell = sheet.cell(row=0, column=column)
            cell.value = t(title)
            cell.style.font.bold = True
            cell.style.alignment.wrap_text = True
            letter = get_column_letter(column+1)
            if title == 'report_timeline_measure':
                sheet.column_dimensions[letter].width = len(cell.value)+50
            else:
                sheet.column_dimensions[letter].width = len(cell.value)+5

        for (row, (module, risk, measure)) in \
                enumerate(self.get_measures(), 1):
            column = 0

            if not getattr(risk, 'is_custom_risk', None):
                zodb_node = self.request.survey.restrictedTraverse(
                    risk.zodb_path.split('/'))
            else:
                zodb_node = None

            for (type, key, title) in self.columns+self.extra_cols:
                value = None
                if type == 'measure':
                    value = getattr(measure, key, None)
                elif type == 'risk':
                    value = getattr(risk, key, None)
                    if key == 'priority':
                        value = self.priority_name(value)
                    elif key == 'title':
                        if zodb_node is None:
                            value = getattr(risk, key, None)
                        elif zodb_node.problem_description and \
                                zodb_node.problem_description.strip():
                            value = zodb_node.problem_description

                elif type == 'module':
                    if key == 'title' and module.depth > 1:
                        titles = []
                        m = module
                        while m:
                            title = getattr(m, 'title', None)
                            if title:
                                titles.append(m.title)
                            m = m.parent
                        titles.reverse()
                        value = ', '.join(titles)
                    else:
                        if module.zodb_path == 'custom-risks':
                            lang = getattr(self.request, 'LANGUAGE', 'en')
                            if "-" in lang:
                                elems = lang.split("-")
                                lang = "{0}_{1}".format(elems[0], elems[1].upper())
                            value = translate(_(
                                'title_other_risks', default=u'Added risks (by you)'),
                                target_language=lang)
                        else:
                            value = getattr(module, key, None)

                sheet.cell(row=row, column=column)\
                    .style.alignment.wrap_text = True  # style
                if key in self.combine_keys and value is not None:
                    # osha wants to combine action_plan (col 3),
                    # prevention_plan and requirements in one cell
                    if not sheet.cell(row=row, column=2).value:
                        sheet.cell(row=row, column=2).value = u''
                    sheet.cell(row=row, column=2).value += '\r\n'+value
                    continue

                if value is not None:
                    sheet.cell(row=row, column=column).value = value
                column += 1
        return book

    def get_measures(self):
        """Find all data that should be included in the report.

        The data is returned as a list of tuples containing a
        :py:class:`Module <euphorie.client.model.Module>`,
        :py:class:`Risk <euphorie.client.model.Risk>` and
        :py:class:`ActionPlan <euphorie.client.model.ActionPlan>`. Each
        entry in the list will correspond to a row in the generated Excel
        file.

        This implementation differs from Euphorie in its ordering:
        it sorts on risk priority instead of start date.
        """
        query = Session.query(model.Module, model.Risk, model.ActionPlan)\
            .filter(sql.and_(model.Module.session == self.session,
                             model.Module.profile_index > -1))\
            .filter(sql.not_(model.SKIPPED_PARENTS))\
            .filter(sql.or_(model.MODULE_WITH_RISK_OR_TOP5_FILTER,
                            model.RISK_PRESENT_OR_TOP5_FILTER))\
            .join((model.Risk,
                   sql.and_(model.Risk.path.startswith(model.Module.path),
                            model.Risk.depth == model.Module.depth+1,
                            model.Risk.session == self.session)))\
            .join((model.ActionPlan,
                   model.ActionPlan.risk_id == model.Risk.id))\
            .order_by(
                sql.case(
                    value=model.Risk.priority,
                    whens={'high': 0, 'medium': 1},
                    else_=2),
                model.Risk.path)
        return [t for t in query.all() if (
            t[-1].planning_start is not None or
            t[-1].planning_end is not None or
            t[-1].responsible is not None or
            t[-1].prevention_plan is not None or
            t[-1].requirements is not None or
            t[-1].budget is not None or
            t[-1].action_plan is not None
        )]


def node_title(node, zodbnode):
    # 2885: Non-present risks and unanswered risks are shown affirmatively,
    # i.e 'title'
    if node.type != "risk" or node.identification in [u"n/a", u"yes", None]:
        return node.title
    # The other two groups of risks are shown negatively, i.e
    # 'problem_description'
    if zodbnode.problem_description and zodbnode.problem_description.strip():
        return zodbnode.problem_description
    return node.title


class OSHAIdentificationReport(report.IdentificationReport):
    """
    Overrides the original IdentificationReport in euphorie.client.survey.py
    in order to provide a new template.

    Please refer to original for more details.
    """
    grok.layer(IOSHAIdentificationPhaseSkinLayer)
    grok.template("report_identification")
    download = False

    def title(self, node, zodbnode):
        return node.title

    def publishTraverse(self, request, name):
        """Check if the user wants to download this report by checking for a
        ``download`` URL entry. This uses a little trick: browser views
        implement `IPublishTraverse`, which allows us to catch traversal steps.
        """
        if name == "download":
            return OSHAIdentificationReportDownload(
                aq_inner(self.context), request)
        else:
            raise NotFound(self, name, request)

    def update(self):
        if survey.redirectOnSurveyUpdate(self.request):
            return

        # 3813: Include children from optional modules.
        # Removed this: .filter(sql.not_(model.SKIPPED_PARENTS))\
        session = Session()
        dbsession = SessionManager.session
        query = session.query(model.SurveyTreeItem)\
            .filter(model.SurveyTreeItem.session == dbsession)\
            .order_by(model.SurveyTreeItem.path)
        self.nodes = query.all()


class OSHAActionPlanReportDownload(report.ActionPlanReportDownload):
    """ Generate and download action report.
    """
    grok.layer(IOSHAReportPhaseSkinLayer)
    grok.name("download")
    download = True

    def update(self):
        """ Fetches the different kinds of risks we are interested in.

            Actioned Nodes
            --------------
            Title: "Risks that have been identified, evaluated and have an
            Action Plan"

            Unactioned Nodes
            ----------------
            Title: "Risks that have been identified but do NOT have an Action
            Plan"

            Unanswered Nodes
            ----------------
            Title: "Hazards/problems that have been "parked" and are still
            to be dealt with"

            Risk not present nodes
            ----------------------
            Title: "Hazards/problems that have been managed or are not present
            in your organisation"
        """
        if survey.redirectOnSurveyUpdate(self.request):
            return

        super(OSHAActionPlanReportDownload, self).update()
        # Returns all identified nodes, with or without action plans
        self.nodes = self.getNodes()

        # Get the extra attributes as per #1517, #1518:
        self.actioned_nodes = utils.get_actioned_nodes(self.nodes)

        self.unactioned_nodes = utils.get_unactioned_nodes(self.nodes)

        self.unanswered_nodes = utils.get_unanswered_nodes(self.session)
        risk_not_present_nodes = utils.get_risk_not_present_nodes(self.session)
        # From the non-present risks, filter out risks from the (un-)/actioned
        # categories. A "priority" risk will always appear in the action plan,
        # even if it has been answered with "Yes"
        self.risk_not_present_nodes = [
            n for n in risk_not_present_nodes if
            n not in self.actioned_nodes and n not in self.unactioned_nodes
        ]

        lang = getattr(self.request, 'LANGUAGE', 'en')
        if "-" in lang:
            elems = lang.split("-")
            lang = "{0}_{1}".format(elems[0], elems[1].upper())
        self.title_custom_risks = translate(_(
            'title_other_risks', default=u'Added risks (by you)'),
            target_language=lang)

    def addReportNodes(self, document, nodes, heading, toc, body):
        """ """
        t = lambda txt: "".join([
            "\u%s?" % str(ord(e)) for e in translate(txt, context=self.request)
        ])
        ss = document.StyleSheet
        toc_props = ParagraphPropertySet()
        toc_props.SetLeftIndent(TabPropertySet.DEFAULT_WIDTH * 1)
        toc_props.SetRightIndent(TabPropertySet.DEFAULT_WIDTH * 1)
        p = Paragraph(ss.ParagraphStyles.Heading6, toc_props)
        p.append(character.Text(heading, TextPropertySet(italic=True)))
        toc.append(p)

        body.append(Paragraph(ss.ParagraphStyles.Heading1, heading))

        survey = self.request.survey
        styles = ss.ParagraphStyles
        header_styles = {
            0: styles.Heading2,
            1: styles.Heading3,
            2: styles.Heading4,
            3: styles.Heading5,
            4: styles.Heading6,
        }
        for node in nodes:
            zodb_node = None
            if node.zodb_path == 'custom-risks':
                title = self.title_custom_risks
            elif getattr(node, 'is_custom_risk', None):
                title = node.title
            else:
                zodb_node = survey.restrictedTraverse(node.zodb_path.split("/"))
                title = node_title(node, zodb_node)

            thin_edge = BorderPropertySet(
                width=20, style=BorderPropertySet.SINGLE)

            if node.depth == 1:
                p = Paragraph(
                    header_styles.get(node.depth, styles.Heading6),
                    FramePropertySet(
                        thin_edge, thin_edge, thin_edge, thin_edge),
                    u"%s %s" % (node.number, title))
            else:
                p = Paragraph(
                    header_styles.get(node.depth, styles.Heading6),
                    u"%s %s" % (node.number, title))
            body.append(p)

            if node.type != "risk":
                continue

            if node.priority:
                if node.priority == "low":
                    level = _("risk_priority_low", default=u"low")
                elif node.priority == "medium":
                    level = _("risk_priority_medium", default=u"medium")
                elif node.priority == "high":
                    level = _("risk_priority_high", default=u"high")

                msg = _("risk_priority",
                        default="This is a ${priority_value} priority risk.",
                        mapping={'priority_value': level})
                body.append(Paragraph(
                    styles.RiskPriority,
                    t(msg)
                ))

            if getattr(node, 'identification', None) == 'no':
                if zodb_node is None:
                    description = node.title
                else:
                    description = zodb_node.description

                body.append(
                    Paragraph(
                        styles.Normal,
                        ParagraphPropertySet(
                            left_indent=300, right_indent=300),
                        t(_(utils.html_unescape(
                            htmllaundry.StripMarkup(description))))
                    ))
                body.append(Paragraph(""))

            if node.comment and node.comment.strip():
                body.append(Paragraph(styles.Comment, node.comment))

            for (idx, measure) in enumerate(node.action_plans):
                if not measure.action_plan:
                    continue

                if len(node.action_plans) == 1:
                    heading = t(_("header_measure_single", default=u"Measure"))
                else:
                    heading = t(
                        _("header_measure",
                            default=u"Measure ${index}",
                            mapping={"index": idx + 1}))

                self.addMeasure(document, heading, body, measure)

    def addMeasure(self, document, heading, section, measure):
        """ Requirements for how the measure section should be displayed are
            in #2611
        """
        t = lambda txt: "".join([
            "\u%s?" % str(ord(e)) for e in translate(txt, context=self.request)
        ])
        ss = document.StyleSheet
        styles = ss.ParagraphStyles

        table = Table(9500)
        thin_edge = BorderPropertySet(width=20, style=BorderPropertySet.SINGLE)
        no_edge = BorderPropertySet(width=0, colour=ss.Colours.White)
        p = Paragraph(
            styles.MeasureHeading,
            ParagraphPropertySet(left_indent=300, right_indent=300),
            t(_("header_measure_single", default=u"Measure")))
        c = Cell(p, FramePropertySet(thin_edge, thin_edge, no_edge, thin_edge))
        table.AddRow(c)

        ss = document.StyleSheet
        styles = document.StyleSheet.ParagraphStyles
        headings = [
            t(_("label_measure_action_plan", default=u"General approach (to "
                u"eliminate or reduce the risk)")),
            t(_("label_measure_prevention_plan", default=u"Specific action(s) "
                u"required to implement this approach")),
            t(_("label_measure_requirements", default=u"Level of expertise "
                u"and/or requirements needed")),
            t(_("label_action_plan_responsible", default=u"Who is "
                u"responsible?")),
            t(_("label_action_plan_budget", default=u"Budget")),
            t(_("label_action_plan_start", default=u"Planning start")),
            t(_("label_action_plan_end", default=u"Planning end")),
        ]
        m = measure
        values = [
            m.action_plan,
            m.prevention_plan,
            m.requirements,
            m.responsible,
            m.budget and str(m.budget) or '',
            m.planning_start and formatDate(self.request, m.planning_start) or '',
            m.planning_end and formatDate(self.request, m.planning_end) or '',
        ]
        for heading, value in zip(headings, values):
            p = Paragraph(
                styles.MeasureField,
                heading
            )
            c = Cell(p, FramePropertySet(no_edge, thin_edge, no_edge, thin_edge))
            table.AddRow(c)

            if headings.index(heading) == len(headings) - 1:
                frame = FramePropertySet(no_edge, thin_edge, thin_edge, thin_edge)
            else:
                frame = FramePropertySet(no_edge, thin_edge, no_edge, thin_edge)

            p = Paragraph(
                styles.Normal,
                ParagraphPropertySet(left_indent=600, right_indent=600),
                value)
            c = Cell(p, frame)
            table.AddRow(c)

        section.append(table)

    def addConsultationBox(self, section, document):
        """ Add the consultation box that needs to be signed by the employer
            and workers.
        """
        ss = document.StyleSheet
        styles = document.StyleSheet.ParagraphStyles
        thin_edge = BorderPropertySet(width=20, style=BorderPropertySet.SINGLE)
        t = lambda txt: "".join([
            "\u%s?" % str(ord(e)) for e in translate(txt, context=self.request)
        ])

        table = Table(9500)
        thin_edge = BorderPropertySet(width=20, style=BorderPropertySet.SINGLE)
        no_edge = BorderPropertySet(width=0, colour=ss.Colours.White)
        p = Paragraph(
            styles.Heading3,
            ParagraphPropertySet(alignment=ParagraphPropertySet.CENTER),
            t(_("header_oira_report_consultation",
                default="Consultation of workers"))
        )
        c = Cell(p, FramePropertySet(thin_edge, thin_edge, no_edge, thin_edge))
        table.AddRow(c)

        p = Paragraph(
            styles.Normal,
            ParagraphPropertySet(alignment=ParagraphPropertySet.LEFT),
            t(_("paragraph_oira_consultation_of_workers",
                default="The undersigned hereby declare that the workers "
                        "have been consulted on the content of this "
                        "document.")),
            LINE
        )
        c = Cell(p, FramePropertySet(no_edge, thin_edge, no_edge, thin_edge))
        table.AddRow(c)

        p = Paragraph(
            styles.Normal,
            ParagraphPropertySet(alignment=ParagraphPropertySet.LEFT),
        )
        employer = t(_("oira_consultation_employer",
                       default="On behalf of the employer:"))
        workers = t(_("oira_consultation_workers",
                    default="On behalf of the workers:"))

        p.append(employer, TAB, TAB, TAB, TAB, workers, LINE, LINE)
        c = Cell(p, FramePropertySet(no_edge, thin_edge, no_edge, thin_edge))
        table.AddRow(c)

        p = Paragraph(
            ParagraphPropertySet(alignment=ParagraphPropertySet.LEFT),
            t(_("oira_survey_date", default="Date:")),
            LINE, LINE
        )
        c = Cell(p, FramePropertySet(no_edge, thin_edge, thin_edge, thin_edge))
        table.AddRow(c)
        section.append(table)

    def render(self):
        """ Mostly a copy of the render method in euphorie.client, but with
            some changes to also show unanswered risks and non-present risks.
            #1517 and #1518
        """
        document = report.createDocument(self.session)
        ss = document.StyleSheet

        # Define some more custom styles
        ss.ParagraphStyles.append(
            ParagraphStyle(
                "RiskPriority",
                TextStyle(
                    TextPropertySet(
                        font=ss.Fonts.Arial,
                        size=22,
                        italic=True,
                        colour=ss.Colours.Blue)),
                ParagraphPropertySet(left_indent=300, right_indent=300))
        )
        ss.ParagraphStyles.append(
            ParagraphStyle(
                "MeasureField",
                TextStyle(
                    TextPropertySet(
                        font=ss.Fonts.Arial,
                        size=18,
                        underline=True)),
                ParagraphPropertySet(left_indent=300, right_indent=300))
        )
        # XXX: This part is removed
        # self.addActionPlan(document)

        # XXX: and replaced with this part:
        t = lambda txt: "".join([
            "\u%s?" % str(ord(e)) for e in translate(txt, context=self.request)
        ])
        toc = createSection(document, self.context, self.request)

        body = Section()
        heading = t(_("header_oira_report_download",
                    default=u"OiRA Report: \"${title}\"",
                    mapping=dict(title=self.session.title)))

        toc.append(Paragraph(
            ss.ParagraphStyles.Heading1,
            ParagraphPropertySet(alignment=ParagraphPropertySet.CENTER),
            heading,
        ))

        if self.session.report_comment:
            # Add comment. #5985
            normal_style = document.StyleSheet.ParagraphStyles.Normal
            toc.append(Paragraph(normal_style, self.session.report_comment))

        toc_props = ParagraphPropertySet()
        toc_props.SetLeftIndent(TabPropertySet.DEFAULT_WIDTH * 1)
        toc_props.SetRightIndent(TabPropertySet.DEFAULT_WIDTH * 1)
        p = Paragraph(ss.ParagraphStyles.Heading6, toc_props)
        txt = t(_("toc_header", default=u"Contents"))
        p.append(character.Text(txt))
        toc.append(p)

        headings = [
            t(_("header_present_risks",
                default=u"Risks that have been identified, "
                        u"evaluated and have an Action Plan")),
            t(_("header_unevaluated_risks",
                default=u"Risks that have been identified but "
                        u"do NOT have an Action Plan")),
            t(_("header_unanswered_risks",
                default=u'Hazards/problems that have been "parked" '
                        u'and are still to be dealt with')),
            t(_("header_risks_not_present",
                default=u"Hazards/problems that have been managed "
                        u"or are not present in your organisation"))
        ]
        nodes = [
            self.actioned_nodes,
            self.unactioned_nodes,
            self.unanswered_nodes,
            self.risk_not_present_nodes,
        ]

        for nodes, heading in zip(nodes, headings):
            if not nodes:
                continue
            self.addReportNodes(document, nodes, heading, toc, body)

        toc.append(Paragraph(LINE))
        body.append(Paragraph(LINE))
        self.addConsultationBox(body, document)
        document.Sections.append(body)
        # Until here...

        renderer = Renderer()
        output = StringIO()
        renderer.Write(document, output)

        filename = translate(
            _("filename_report_actionplan",
                default=u"Action plan ${title}",
                mapping=dict(title=self.session.title)),
            context=self.request,)
        self.request.response.setHeader(
            "Content-Disposition",
            "attachment; filename=\"%s.rtf\"" % filename.encode("utf-8"))
        self.request.response.setHeader("Content-Type", "application/rtf")
        return output.getvalue()


class OSHAIdentificationReportDownload(report.IdentificationReportDownload):
    """Generate identification report in RTF form.
    """
    grok.layer(IOSHAIdentificationPhaseSkinLayer)

    def getNodes(self):
        """ Return an ordered list of all relevant tree items for the current
            survey.
        """
        # 3813: Include children from optional modules.
        # Removed this: .filter(sql.not_(model.SKIPPED_PARENTS))\
        query = Session.query(model.SurveyTreeItem)\
            .filter(model.SurveyTreeItem.session == self.session)\
            .order_by(model.SurveyTreeItem.path)
        return query.all()

    def addIdentificationResults(self, document):
        survey = self.request.survey
        section = createIdentificationReportSection(
            document, self.context, self.request)

        styles = document.StyleSheet.ParagraphStyles
        header_styles = {
            0: styles.Heading2,
            1: styles.Heading3,
            2: styles.Heading4,
            3: styles.Heading5,
            4: styles.Heading6,
        }

        for node in self.getNodes():
            section.append(
                Paragraph(
                    header_styles.get(node.depth, styles.Heading6),
                    u"%s %s" % (node.number, node.title))
            )

            if node.type != "risk":
                continue

            if getattr(node, 'is_custom_node', None):
                description = survey.restrictedTraverse(
                        node.zodb_path.split("/")).description
            else:
                description = node.title

            section.append(
                Paragraph(
                    styles.Normal,
                    utils.html_unescape(
                        htmllaundry.StripMarkup(description))
                )
            )

            for i in range(0, 8):
                p = Paragraph(styles.Normal, " ")
                section.append(p)

            tabs = TabPropertySet(
                section.TwipsToRightMargin(),
                alignment=TabPropertySet.RIGHT,
                leader=getattr(TabPropertySet, 'UNDERLINE')
            )
            p = Paragraph(styles.Normal, ParagraphPropertySet(tabs=[tabs]))
            p.append(TAB)
            section.append(p)

            if node.comment and node.comment.strip():
                section.append(Paragraph(styles.Comment, node.comment))


def createIdentificationReportSection(document, survey, request):
    t = lambda txt: "".join([
        "\u%s?" % str(ord(e)) for e in translate(txt, context=request)
    ])
    section = Section()

    footer_txt = t(
        _("report_identification_revision",
            default=u"This document was based on the OiRA Tool '${title}' of "
                    u"revision date ${date}.",
            mapping={"title": survey.published[1],
                    "date": formatDate(request, survey.published[2])}))
    header = Table(4750, 4750)
    c1 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        SessionManager.session.title))

    pp = ParagraphPropertySet
    header_props = pp(alignment=pp.RIGHT)
    c2 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        header_props,
        formatDate(request, datetime.today())))
    header.AddRow(c1, c2)
    section.Header.append(header)

    footer = Table(9000, 500)
    c1 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        pp(alignment=pp.LEFT),
        footer_txt))
    c2 = Cell(Paragraph(pp(alignment=pp.RIGHT), PAGE_NUMBER))
    footer.AddRow(c1, c2)
    section.Footer.append(footer)
    section.SetBreakType(section.PAGE)
    document.Sections.append(section)
    return section


def createSection(document, survey, request):
    t = lambda txt: "".join([
        "\u%s?" % str(ord(e)) for e in translate(txt, context=request)
    ])
    section = Section(break_type=Section.PAGE, first_page_number=1)
    footer_txt = t(
        _("report_survey_revision",
            default=u"This report was based on the OiRA Tool '${title}' "\
                    u"of revision date ${date}.",
            mapping={"title": survey.published[1],
                    "date": formatDate(request, survey.published[2])}))

    header = Table(4750, 4750)
    c1 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        survey.published[1]))

    pp = ParagraphPropertySet
    header_props = pp(alignment=pp.RIGHT)
    c2 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        header_props,
        formatDate(request, datetime.today())))
    header.AddRow(c1, c2)
    section.Header.append(header)

    footer = Table(9000, 500)
    # rtfng does not like unicode footers
    c1 = Cell(Paragraph(
        document.StyleSheet.ParagraphStyles.Footer,
        pp(alignment=pp.LEFT),
        footer_txt))

    c2 = Cell(Paragraph(pp(alignment=pp.RIGHT), PAGE_NUMBER))
    footer.AddRow(c1, c2)
    section.Footer.append(footer)
    document.Sections.append(section)
    return section


class RisksOverview(OSHAStatus):
    """ Implements the "Overview of Risks" report, see #10967
    """
    grok.context(PathGhost)
    grok.layer(IOSHAReportPhaseSkinLayer)
    grok.template("risks_overview")
    grok.name("risks_overview")

    def is_skipped_from_risk_list(self, r):
        if r['identification'] == 'yes':
            return True


class MeasuresOverview(OSHAStatus):
    """ Implements the "Overview of Measures" report, see #10967
    """
    grok.context(PathGhost)
    grok.layer(IOSHAReportPhaseSkinLayer)
    grok.template("measures_overview")
    grok.require('euphorie.client.ViewSurvey')
    grok.name("measures_overview")

    def update(self):
        self.session = SessionManager.session
        lang = getattr(self.request, 'LANGUAGE', 'en')
        if "-" in lang:
            lang = lang.split("-")[0]
        now = datetime.now()
        next_month = datetime(now.year, (now.month + 1) % 12 or 12, 1)
        month_after_next = datetime(now.year, (now.month + 2) % 12 or 12, 1)
        self.months = []
        self.months.append(now.strftime('%b'))
        self.months.append(next_month.strftime('%b'))
        self.months.append(month_after_next.strftime('%b'))
        self.monthstrings = [
            translate(
                PloneLocalesFactory(
                    "month_{0}_abbr".format(month.lower()),
                    default=month,
                ),
                target_language=lang,
            )
            for month in self.months
        ]

        query = Session.query(model.Module, model.Risk, model.ActionPlan)\
            .filter(sql.and_(model.Module.session == self.session,
                             model.Module.profile_index > -1))\
            .filter(sql.not_(model.SKIPPED_PARENTS))\
            .filter(sql.or_(model.MODULE_WITH_RISK_OR_TOP5_FILTER,
                            model.RISK_PRESENT_OR_TOP5_FILTER))\
            .join((model.Risk,
                   sql.and_(model.Risk.path.startswith(model.Module.path),
                            model.Risk.depth == model.Module.depth+1,
                            model.Risk.session == self.session)))\
            .join((model.ActionPlan,
                   model.ActionPlan.risk_id == model.Risk.id))\
            .order_by(
                sql.case(
                    value=model.Risk.priority,
                    whens={'high': 0, 'medium': 1},
                    else_=2),
                model.Risk.path)
        measures = [t for t in query.all() if (
            (t[-1].planning_start is not None
                and t[-1].planning_start.strftime('%b') in self.months) and
            (
                t[-1].planning_end is not None or
                t[-1].responsible is not None or
                t[-1].prevention_plan is not None or
                t[-1].requirements is not None or
                t[-1].budget is not None or
                t[-1].action_plan is not None
            )
        )]

        modulesdict = defaultdict(lambda: defaultdict(list))
        for module, risk, action in measures:
            if 'custom-risks' not in risk.zodb_path:
                risk_obj = self.request.survey.restrictedTraverse(risk.zodb_path.split('/'))
                title = risk_obj and risk_obj.problem_description or risk.title
            else:
                title = risk.title
            modulesdict[module][risk.priority].append(
                {'title': title,
                 'description': action.action_plan,
                 'months': [action.planning_start and
                            action.planning_start.month == m.month
                            for m in [now, next_month, month_after_next]],
                 })

        # re-use top-level module computation from the Status overview
        modules = self.getModules()
        main_modules = {}
        for module, risks in sorted(modulesdict.items(), key=lambda m: m[0].zodb_path):
            module_obj = self.request.survey.restrictedTraverse(module.zodb_path.split('/'))
            if (
                IProfileQuestion.providedBy(module_obj) or
                ICustomRisksModule.providedBy(module_obj) or
                module.depth >= 3
            ):
                path = module.path[:6]
            else:
                path = module.path[:3]
            if path in main_modules:
                for prio in risks.keys():
                    if prio in main_modules[path]['risks']:
                        main_modules[path]['risks'][prio].extend(risks[prio])
                    else:
                        main_modules[path]['risks'][prio] = risks[prio]
            else:
                title = modules[path]['title']
                number = modules[path]['number']
                main_modules[path] = {'name': title, 'number': number, 'risks': risks}

        self.modules = []
        for key in sorted(main_modules.keys()):
            self.modules.append(main_modules[key])
